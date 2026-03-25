"""
SV Dugout Pulse — High School Stats Integration

Parses a manually-maintained Google Sheet (exported as .xlsx) containing
per-game HS stats.  Each tab is a game date; hitter and pitcher sections
are detected by their header rows.
"""

from __future__ import annotations

import json
import logging
import os
import re
import tempfile
from datetime import date, datetime, timedelta
from typing import Optional

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from .config import (
    HS_GAME_LOG_PATH,
    HS_NAME_ALIASES,
    HS_STATS_XLSX_URL,
)
from .historical_stats import ip_to_outs, outs_to_ip_display

logger = logging.getLogger(__name__)


def _make_http_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


_http = _make_http_session()

# Year for the current HS season
_SEASON_YEAR = 2026


def _parse_tab_date(tab_name: str) -> Optional[date]:
    """Parse a tab name like '23', '210', '324' into a date.

    Digits concatenate M/D with no separator.  Try month=first 1 digit,
    then month=first 2 digits.  Use _SEASON_YEAR.
    """
    digits = tab_name.strip()
    if not digits.isdigit() or len(digits) < 2:
        return None

    # Try 1-digit month first (e.g. "23" -> month=2, day=3)
    m1 = int(digits[0])
    d1_str = digits[1:]
    if 1 <= m1 <= 9 and d1_str.isdigit():
        d1 = int(d1_str)
        if 1 <= d1 <= 31:
            try:
                return date(_SEASON_YEAR, m1, d1)
            except ValueError:
                pass

    # Try 2-digit month (e.g. "1015" -> month=10, day=15)
    if len(digits) >= 3:
        m2 = int(digits[:2])
        d2_str = digits[2:]
        if 1 <= m2 <= 12 and d2_str.isdigit():
            d2 = int(d2_str)
            if 1 <= d2 <= 31:
                try:
                    return date(_SEASON_YEAR, m2, d2)
                except ValueError:
                    pass

    return None


def _normalize_name(raw_name: str) -> str:
    """Strip whitespace, remove (DH) suffix, apply alias mapping."""
    name = raw_name.strip()
    # Remove (DH) or similar parenthetical suffixes
    name = re.sub(r"\s*\(DH\)\s*$", "", name, flags=re.IGNORECASE).strip()

    # Check alias table (check original raw name first for entries like "David Vargas (DH)")
    if raw_name.strip() in HS_NAME_ALIASES:
        return HS_NAME_ALIASES[raw_name.strip()]
    if name in HS_NAME_ALIASES:
        return HS_NAME_ALIASES[name]

    return name


# =========================================================================
# HSSheetParser — downloads and parses the xlsx
# =========================================================================

class HSSheetParser:
    """Download and parse the HS stats Google Sheet (xlsx export)."""

    def __init__(self):
        self._parsed: list[dict] | None = None

    def parse_all(self) -> list[dict]:
        """Download xlsx and parse all tabs.

        Returns list of dicts:
            [{"date": date, "hitters": [...], "pitchers": [...]}, ...]
        """
        if self._parsed is not None:
            return self._parsed

        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed — cannot parse HS stats sheet")
            self._parsed = []
            return self._parsed

        # Download xlsx to a temp file
        tmp_path = None
        try:
            resp = _http.get(HS_STATS_XLSX_URL, timeout=60)
            resp.raise_for_status()
            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            with os.fdopen(fd, "wb") as f:
                f.write(resp.content)
            logger.info("Downloaded HS stats xlsx (%d bytes)", len(resp.content))
        except Exception:
            logger.exception("Failed to download HS stats xlsx")
            self._parsed = []
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)
            return self._parsed

        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            results = []
            for sheet_name in wb.sheetnames:
                game_date = _parse_tab_date(sheet_name)
                if game_date is None:
                    logger.debug("Skipping tab '%s' — cannot parse date", sheet_name)
                    continue
                ws = wb[sheet_name]
                parsed = self._parse_tab(ws, sheet_name)
                parsed["date"] = game_date
                results.append(parsed)
                logger.debug(
                    "Parsed tab '%s' (%s): %d hitters, %d pitchers",
                    sheet_name, game_date, len(parsed["hitters"]), len(parsed["pitchers"]),
                )
            wb.close()
            self._parsed = results
            logger.info("Parsed %d game tabs from HS sheet", len(results))
            return results
        except Exception:
            logger.exception("Failed to parse HS stats xlsx")
            self._parsed = []
            return self._parsed
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    def _parse_tab(self, ws, tab_name: str) -> dict:
        """Parse a single worksheet tab into hitter and pitcher sections."""
        rows = list(ws.iter_rows(values_only=True))

        hitters = []
        pitchers = []

        # Find hitter header row (has PLAYER and AB columns)
        hitter_header_idx = None
        pitcher_header_idx = None
        for i, row in enumerate(rows):
            if row is None:
                continue
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if "PLAYER" in cells and "AB" in cells:
                hitter_header_idx = i
            elif "PLAYER" in cells and "IP" in cells:
                pitcher_header_idx = i

        # Parse hitters
        if hitter_header_idx is not None:
            header = rows[hitter_header_idx]
            col_map = {}
            for ci, cell in enumerate(header):
                if cell is not None:
                    col_map[str(cell).strip().upper()] = ci

            # Determine end of hitter section (blank row or pitcher header)
            end_idx = pitcher_header_idx if pitcher_header_idx is not None else len(rows)
            for i in range(hitter_header_idx + 1, end_idx):
                row = rows[i]
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue
                player_col = col_map.get("PLAYER")
                if player_col is None or player_col >= len(row):
                    continue
                raw_name = row[player_col]
                if raw_name is None or str(raw_name).strip() == "":
                    continue
                name = _normalize_name(str(raw_name))

                def _get(key, default=0):
                    idx = col_map.get(key)
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return default
                    try:
                        return int(row[idx])
                    except (ValueError, TypeError):
                        return default

                # Game result column
                result_col = col_map.get("GAME RESULT")
                game_result = ""
                if result_col is not None and result_col < len(row) and row[result_col]:
                    game_result = str(row[result_col]).strip()

                entry = {
                    "player": name,
                    "ab": _get("AB"),
                    "r": _get("R"),
                    "h": _get("H"),
                    "2b": _get("2B"),
                    "3b": _get("3B"),
                    "hr": _get("HR"),
                    "rbi": _get("RBI"),
                    "bb": _get("BB"),
                    "k": _get("SO"),
                    "game_result": game_result,
                }
                hitters.append(entry)

        # Parse pitchers
        if pitcher_header_idx is not None:
            header = rows[pitcher_header_idx]
            col_map = {}
            for ci, cell in enumerate(header):
                if cell is not None:
                    col_map[str(cell).strip().upper()] = ci

            for i in range(pitcher_header_idx + 1, len(rows)):
                row = rows[i]
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue
                player_col = col_map.get("PLAYER")
                if player_col is None or player_col >= len(row):
                    continue
                raw_name = row[player_col]
                if raw_name is None or str(raw_name).strip() == "":
                    continue
                name = _normalize_name(str(raw_name))

                def _pget(key, default=0):
                    idx = col_map.get(key)
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return default
                    try:
                        return float(row[idx]) if key == "IP" else int(row[idx])
                    except (ValueError, TypeError):
                        return default

                # Game result column
                result_col = col_map.get("GAME RESULT")
                game_result = ""
                if result_col is not None and result_col < len(row) and row[result_col]:
                    game_result = str(row[result_col]).strip()

                # K% column (informational; we compute from raw stats)
                entry = {
                    "player": name,
                    "ip": _pget("IP"),
                    "h": _pget("H"),
                    "r": _pget("R"),
                    "er": _pget("ER"),
                    "bb": _pget("BB"),
                    "k": _pget("SO"),
                    "game_result": game_result,
                }
                pitchers.append(entry)

        return {"hitters": hitters, "pitchers": pitchers}

    def get_all_player_names(self) -> set[str]:
        """Return all normalized player names across all tabs."""
        data = self.parse_all()
        names = set()
        for game in data:
            for h in game["hitters"]:
                names.add(h["player"])
            for p in game["pitchers"]:
                names.add(p["player"])
        return names

    def get_position_for_player(self, name: str) -> str:
        """Determine position: 'Pitcher', 'Two-Way', or 'Hitter'."""
        data = self.parse_all()
        in_hitting = False
        in_pitching = False
        for game in data:
            for h in game["hitters"]:
                if h["player"] == name:
                    in_hitting = True
            for p in game["pitchers"]:
                if p["player"] == name:
                    in_pitching = True
        if in_pitching and in_hitting:
            return "Two-Way"
        if in_pitching:
            return "Pitcher"
        return "Hitter"


# =========================================================================
# HSGameLog — persist per-game stats (mirrors ncaa_game_log.json pattern)
# =========================================================================

class HSGameLog:
    """Load/save HS game log from disk and query by player/date."""

    def __init__(self):
        self._log: dict[str, list] = {}
        if os.path.exists(HS_GAME_LOG_PATH):
            try:
                with open(HS_GAME_LOG_PATH) as f:
                    self._log = json.load(f)
            except Exception:
                logger.error("HS game log corrupted — starting fresh")
                self._log = {}

    def save(self):
        """Write game log to disk atomically."""
        os.makedirs(os.path.dirname(HS_GAME_LOG_PATH), exist_ok=True)
        fd, tmp = tempfile.mkstemp(
            dir=os.path.dirname(HS_GAME_LOG_PATH), suffix=".tmp"
        )
        try:
            with os.fdopen(fd, "w") as f:
                json.dump(self._log, f, indent=2, ensure_ascii=False)
            os.replace(tmp, HS_GAME_LOG_PATH)
        except BaseException:
            os.unlink(tmp)
            raise
        logger.info("Saved HS game log (%d players)", len(self._log))

    def update_from_sheet(self, parsed_data: list[dict]):
        """Merge parsed sheet data into the game log, deduplicating by player+date."""
        added = 0
        for game in parsed_data:
            game_date = game["date"].isoformat()  # YYYY-MM-DD

            for h in game["hitters"]:
                key = h["player"]
                entry = {
                    "date": game_date,
                    "opponent": h.get("game_result", ""),
                    "stats": {
                        "ab": h["ab"],
                        "h": h["h"],
                        "2b": h.get("2b", 0),
                        "3b": h.get("3b", 0),
                        "hr": h.get("hr", 0),
                        "rbi": h.get("rbi", 0),
                        "r": h.get("r", 0),
                        "bb": h.get("bb", 0),
                        "k": h.get("k", 0),
                    },
                    "type": "hitting",
                }
                if self._add_if_new(key, entry):
                    added += 1

            for p in game["pitchers"]:
                key = p["player"]
                # Convert IP to baseball notation string
                ip_val = p["ip"]
                ip_str = str(ip_val)
                # If it's a float like 5.1, keep as-is (already baseball notation)
                # If it's a whole number like 6.0, show as "6"
                if isinstance(ip_val, float) and ip_val == int(ip_val):
                    ip_str = str(int(ip_val))
                elif isinstance(ip_val, (int, float)):
                    ip_str = str(ip_val)

                entry = {
                    "date": game_date,
                    "opponent": p.get("game_result", ""),
                    "stats": {
                        "ip": ip_str,
                        "h": p["h"],
                        "r": p.get("r", 0),
                        "er": p.get("er", 0),
                        "bb": p.get("bb", 0),
                        "k": p.get("k", 0),
                    },
                    "type": "pitching",
                }
                if self._add_if_new(key, entry):
                    added += 1

        if added > 0:
            self.save()
        logger.info("HS game log: merged %d new entries", added)

    def _add_if_new(self, key: str, entry: dict) -> bool:
        """Add entry if not already present for this player+date+type."""
        existing = self._log.get(key, [])
        for e in existing:
            if e["date"] == entry["date"] and e.get("type") == entry.get("type"):
                return False
        self._log.setdefault(key, []).append(entry)
        return True

    def get_player_stats_for_date(
        self, player_name: str, target_date: date
    ) -> list[dict]:
        """Return game log entries for a player on a specific date."""
        entries = self._log.get(player_name, [])
        target_str = target_date.isoformat()
        return [e for e in entries if e["date"] == target_str]

    def get_window_stats(
        self,
        player_name: str,
        position: str,
        start_date: date,
        end_date: date,
    ) -> tuple[Optional[dict], list]:
        """Aggregate stats across a date range.

        Returns (stats_dict, game_entries) matching the format used by
        NCAAGameLogAggregator.
        """
        entries = self._log.get(player_name, [])

        in_range = []
        for e in entries:
            try:
                d = datetime.strptime(e["date"], "%Y-%m-%d").date()
                if start_date <= d <= end_date:
                    in_range.append(e)
            except (ValueError, KeyError):
                continue

        if not in_range:
            return None, []

        # Determine pitcher vs hitter
        is_pitcher = position == "Pitcher"
        if position == "Two-Way":
            pitcher_entries = [
                e for e in in_range if e.get("type") == "pitching"
            ]
            is_pitcher = len(pitcher_entries) > len(in_range) - len(pitcher_entries)

        if is_pitcher:
            pitching = [e for e in in_range if e.get("type") == "pitching"]
            if not pitching:
                return None, []
            return self._aggregate_pitcher(pitching)
        else:
            hitting = [e for e in in_range if e.get("type") == "hitting"]
            if not hitting:
                return None, []
            return self._aggregate_hitter(hitting)

    def _aggregate_hitter(self, entries: list[dict]) -> tuple[dict, list]:
        totals = {
            "h": 0, "ab": 0, "hr": 0, "2b": 0, "3b": 0,
            "rbi": 0, "r": 0, "bb": 0, "k": 0,
        }
        game_entries = []

        for e in entries:
            s = e.get("stats", {})
            totals["h"] += int(s.get("h", 0))
            totals["ab"] += int(s.get("ab", 0))
            totals["hr"] += int(s.get("hr", 0))
            totals["2b"] += int(s.get("2b", 0))
            totals["3b"] += int(s.get("3b", 0))
            totals["rbi"] += int(s.get("rbi", 0))
            totals["r"] += int(s.get("r", 0))
            totals["bb"] += int(s.get("bb", 0))
            totals["k"] += int(s.get("k", 0))
            game_entries.append({
                "date": e["date"],
                "opponent": e.get("opponent", ""),
                "stats": {
                    "h": int(s.get("h", 0)),
                    "ab": int(s.get("ab", 0)),
                    "hr": int(s.get("hr", 0)),
                    "rbi": int(s.get("rbi", 0)),
                    "r": int(s.get("r", 0)),
                    "bb": int(s.get("bb", 0)),
                    "k": int(s.get("k", 0)),
                    "sb": 0,
                    "hbp": 0,
                },
            })

        game_entries.sort(key=lambda g: g["date"], reverse=True)

        pa = totals["ab"] + totals["bb"]
        avg = totals["h"] / totals["ab"] if totals["ab"] > 0 else 0
        obp = (totals["h"] + totals["bb"]) / pa if pa > 0 else 0
        singles = totals["h"] - totals["2b"] - totals["3b"] - totals["hr"]
        tb = singles + (2 * totals["2b"]) + (3 * totals["3b"]) + (4 * totals["hr"])
        slg = tb / totals["ab"] if totals["ab"] > 0 else 0
        ops = obp + slg
        k_pct = totals["k"] / pa if pa > 0 else 0
        bb_pct = totals["bb"] / pa if pa > 0 else 0

        stats = {
            "games_played": len(entries),
            "pa": pa, "ab": totals["ab"], "h": totals["h"],
            "hr": totals["hr"], "rbi": totals["rbi"], "r": totals["r"],
            "bb": totals["bb"], "k": totals["k"], "sb": 0,
            "avg": avg, "obp": obp, "slg": slg, "ops": ops,
            "k_pct": k_pct, "bb_pct": bb_pct,
            "is_pitcher": False,
        }
        return stats, game_entries

    def _aggregate_pitcher(self, entries: list[dict]) -> tuple[dict, list]:
        total_outs = 0
        totals = {"er": 0, "k": 0, "bb": 0, "h": 0}
        game_entries = []

        for e in entries:
            s = e.get("stats", {})
            total_outs += ip_to_outs(s.get("ip", "0"))
            totals["er"] += int(s.get("er", 0))
            totals["k"] += int(s.get("k", 0))
            totals["bb"] += int(s.get("bb", 0))
            totals["h"] += int(s.get("h", 0))
            game_entries.append({
                "date": e["date"],
                "opponent": e.get("opponent", ""),
                "stats": {
                    "ip": str(s.get("ip", "0")),
                    "er": int(s.get("er", 0)),
                    "k": int(s.get("k", 0)),
                    "bb": int(s.get("bb", 0)),
                    "h": int(s.get("h", 0)),
                },
            })

        game_entries.sort(key=lambda g: g["date"], reverse=True)

        ip = total_outs / 3
        era = (totals["er"] * 9) / ip if ip > 0 else 0
        whip = (totals["bb"] + totals["h"]) / ip if ip > 0 else 0
        k_per_9 = (totals["k"] * 9) / ip if ip > 0 else 0
        bb_per_9 = (totals["bb"] * 9) / ip if ip > 0 else 0
        bf = total_outs + totals["h"] + totals["bb"]
        k_pct = totals["k"] / bf if bf > 0 else 0
        bb_pct = totals["bb"] / bf if bf > 0 else 0

        stats = {
            "games_played": len(entries),
            "ip": ip,
            "ip_display": outs_to_ip_display(total_outs),
            "h": totals["h"], "er": totals["er"], "bb": totals["bb"],
            "k": totals["k"], "hr": 0,
            "w": 0, "l": 0, "sv": 0,
            "era": era, "whip": whip,
            "k_per_9": k_per_9, "bb_per_9": bb_per_9,
            "k_pct": k_pct, "bb_pct": bb_pct,
            "is_pitcher": True,
        }
        return stats, game_entries


# =========================================================================
# HSStatsFetcher — unified interface matching Pro/NCAA fetchers
# =========================================================================

class HSStatsFetcher:
    """Fetch HS stats from the Google Sheet.

    Matches the interface of ProStatsFetcher / NCAAStatsFetcher.
    """

    def __init__(self):
        self._parser = HSSheetParser()
        self._game_log = HSGameLog()
        self._initialized = False

    def _ensure_initialized(self):
        if self._initialized:
            return
        try:
            parsed = self._parser.parse_all()
            self._game_log.update_from_sheet(parsed)
            self._initialized = True
        except Exception:
            logger.exception("Failed to initialize HS stats — will return empty stats")
            self._initialized = True  # Don't retry on every call

    @property
    def parser(self) -> HSSheetParser:
        self._ensure_initialized()
        return self._parser

    @property
    def game_log(self) -> HSGameLog:
        self._ensure_initialized()
        return self._game_log

    def _build_stats_from_entries(
        self, player: dict, entries: list[dict], game_date: date
    ) -> dict:
        """Build a stats dict from game log entries for a specific date."""
        from .stats_engine import empty_stats

        if not entries:
            result = empty_stats()
            result["data_source"] = "HS Sheet"
            result["stats_summary"] = "No game today"
            return result

        name = player.get("player_name", "")
        position = player.get("position", "Hitter")

        # Pick the appropriate entry type based on position
        hitting_entries = [e for e in entries if e.get("type") == "hitting"]
        pitching_entries = [e for e in entries if e.get("type") == "pitching"]

        is_pitcher = position == "Pitcher"
        if position == "Two-Way":
            is_pitcher = len(pitching_entries) > len(hitting_entries)

        if is_pitcher and pitching_entries:
            e = pitching_entries[0]
            s = e.get("stats", {})
            ip = s.get("ip", "0")
            k = int(s.get("k", 0))
            bb = int(s.get("bb", 0))
            h = int(s.get("h", 0))
            er = int(s.get("er", 0))
            r = int(s.get("r", 0))

            return {
                "stats_summary": f"{ip} IP, {k} K, {er} ER",
                "game_context": e.get("opponent", ""),
                "game_status": "Final",
                "game_time": None,
                "game_date": game_date.isoformat(),
                "next_game": None,
                "is_pitcher_line": True,
                "ip": float(ip) if ip else 0.0,
                "earned_runs": er,
                "strikeouts": k,
                "walks_allowed": bb,
                "hits_allowed": h,
                "saves": 0,
                "win": False,
                "loss": False,
                "quality_start": False,
                "is_debut": False,
                "data_source": "HS Sheet",
                "hits": 0, "at_bats": 0, "home_runs": 0, "rbi": 0,
                "runs": 0, "stolen_bases": 0, "doubles": 0, "triples": 0,
                "hit_by_pitch": 0,
            }
        elif hitting_entries:
            e = hitting_entries[0]
            s = e.get("stats", {})
            ab = int(s.get("ab", 0))
            h = int(s.get("h", 0))
            hr = int(s.get("hr", 0))
            rbi = int(s.get("rbi", 0))
            r = int(s.get("r", 0))
            bb = int(s.get("bb", 0))
            k = int(s.get("k", 0))
            doubles = int(s.get("2b", 0))
            triples = int(s.get("3b", 0))

            parts = []
            parts.append(f"{h}-{ab}")
            if hr:
                parts.append(f"{hr} HR")
            if rbi:
                parts.append(f"{rbi} RBI")
            if r:
                parts.append(f"{r} R")
            if bb:
                parts.append(f"{bb} BB")

            return {
                "stats_summary": ", ".join(parts),
                "game_context": e.get("opponent", ""),
                "game_status": "Final",
                "game_time": None,
                "game_date": game_date.isoformat(),
                "next_game": None,
                "is_pitcher_line": False,
                "hits": h, "at_bats": ab, "home_runs": hr,
                "rbi": rbi, "runs": r, "stolen_bases": 0,
                "doubles": doubles, "triples": triples,
                "hit_by_pitch": 0,
                "walks": bb, "strikeouts": k,
                "ip": 0.0, "earned_runs": 0, "walks_allowed": 0,
                "hits_allowed": 0, "saves": 0,
                "win": False, "loss": False,
                "quality_start": False, "is_debut": False,
                "data_source": "HS Sheet",
            }

        result = empty_stats()
        result["data_source"] = "HS Sheet"
        return result

    def fetch(self, player: dict) -> dict:
        """Fetch stats for today's game (if any)."""
        from .stats_engine import empty_stats

        self._ensure_initialized()
        name = player.get("player_name", "")
        today = date.today()

        try:
            entries = self._game_log.get_player_stats_for_date(name, today)
            return self._build_stats_from_entries(player, entries, today)
        except Exception:
            logger.exception("HS fetch failed for %s", name)
            result = empty_stats()
            result["data_source"] = "HS Sheet"
            return result

    def fetch_all(self, player: dict) -> list[dict]:
        """Fetch all stats for today (supports multiple entries)."""
        return [self.fetch(player)]

    def fetch_yesterday(self, player: dict) -> Optional[dict]:
        """Fetch yesterday's stats."""
        from .stats_engine import empty_stats

        self._ensure_initialized()
        name = player.get("player_name", "")
        yesterday = date.today() - timedelta(days=1)

        try:
            entries = self._game_log.get_player_stats_for_date(name, yesterday)
            if not entries:
                return None
            result = self._build_stats_from_entries(player, entries, yesterday)
            result["is_yesterday"] = True
            return result
        except Exception:
            logger.exception("HS fetch_yesterday failed for %s", name)
            return None

    def fetch_all_yesterday(self, player: dict) -> list[dict]:
        """Fetch yesterday's stats as a list."""
        result = self.fetch_yesterday(player)
        return [result] if result else []
